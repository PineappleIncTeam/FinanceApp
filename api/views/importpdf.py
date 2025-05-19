from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import IsAuthenticated
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font

from api.models import Operation, Category
from api.serializers import OperationSerializer


class OperationPDFView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OperationSerializer

    def get_queryset(self):
        queryset = Operation.objects.filter(user=self.request.user)

        operation_type = self.request.query_params.get('type')

        if operation_type:
            queryset = queryset.filter(type=operation_type)
        return queryset.select_related('categories')

    @swagger_auto_schema(
        operation_id='Получение выписки в формате pdf',
        operation_description='Получение .pdf файла',
        manual_parameters=[
            openapi.Parameter(
                name="type",
                in_=openapi.IN_QUERY,
                description="Фильтр по типу операций",
                type=openapi.TYPE_STRING,
                enum=["targets", "outcome", "income"]
            ),
        ],
        responses = {
            200: openapi.Response(description="Файл успешно получен", schema=OperationSerializer),
    })
    def get(self, request):
        # try:
        #     pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))  # Для Windows
        # except:
        #     try:
        #         pdfmetrics.registerFont(TTFont('DejaVuSans', 'DejaVuSans.ttf'))  # Для Linux
        #     except:
        #         # Можно использовать встроенный шрифт, если другие не доступны
        # pdfmetrics.registerFont(TTFont('Pdfa', 'pdfa.ttf'))
        pdfmetrics.registerFont(TTFont('Arial', 'arial.ttf'))
        operations = []
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        for operation in serializer.data:
            operations.append({
                "category": Category.objects.get(id=operation["categories"]).name,
                "amount": operation["amount"],
                "date": operation["date"]
            })
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                    leftMargin=2 * cm, rightMargin=2 * cm,
                                    topMargin=2 * cm, bottomMargin=2 * cm,
                                    encoding='utf-8')

            elements = []
            styles = getSampleStyleSheet()
            font_name = 'Arial'

            styles['Title'].fontName = font_name
            styles['Normal'].fontName = font_name

            title = Paragraph("Freenance", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))

            # Подготовка данных для таблицы
            table_data = [
                ["Date", "Category", "Amount"]  # Заголовки столбцов
            ]

            for op in operations:
                table_data.append([
                    op['date'],
                    op['category'],
                    str(op['amount'])
                ])
            total = sum(float(op['amount']) for op in operations)
            table_data.append([
                "",
                "Total",
                str(total)
            ])
            table = Table(table_data, colWidths=[6 * cm, 4 * cm, 4 * cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3A5FCD')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), font_name),
                ('FONTNAME', (0, 1), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))

            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            pdf_data = buffer.getvalue()
            if len(pdf_data) < 100:
                raise ValueError("Generated PDF is empty")

            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="freenance_report.pdf"'
            return response

        except Exception as e:
            return HttpResponse(
                f"Error generating PDF: {str(e)}",
                status=500,
                content_type='text/plain'
            )


class OperationXLSView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OperationSerializer

    def get_queryset(self):
        queryset = Operation.objects.filter(user=self.request.user)

        operation_type = self.request.query_params.get('type')

        if operation_type:
            queryset = queryset.filter(type=operation_type)
        return queryset.select_related('categories')


    @swagger_auto_schema(
        operation_id='Получение выписки в формате xls',
        operation_description='Получение .xls файла',
        manual_parameters=[
            openapi.Parameter(
                name="type",
                in_=openapi.IN_QUERY,
                description="Фильтр по типу операций",
                type=openapi.TYPE_STRING,
                enum=["targets", "outcome", "income"]
            ),
        ],
        responses = {
            200: openapi.Response(description="Файл успешно получен", schema=OperationSerializer),
    })
    def get(self, request):
        operations = []
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        for operation in serializer.data:
            operations.append({
                "category": Category.objects.get(id=operation["categories"]).name,
                "amount": operation["amount"],
                "date": operation["date"]
            })

        try:
            buffer = BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = "Freenance"

            ws['A1'] = "Freenance report"
            ws['A1'].font = Font(bold=True, size=16)

            headers = ["Date", "Category", "Amount"]
            ws.append(headers)

            for cell in ws[2]:
                cell.font = Font(bold=True)

            for op in operations:
                ws.append([op['date'], op['category'], op['amount']])

            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(buffer)
            buffer.seek(0)

            if buffer.getbuffer().nbytes < 100:
                raise ValueError("Generated Excel file is empty")

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="financial_report.xlsx"'
            return response

        except Exception as e:
            return HttpResponse(
                f"Error generating Excel: {str(e)}",
                status=500,
                content_type='text/plain'
            )