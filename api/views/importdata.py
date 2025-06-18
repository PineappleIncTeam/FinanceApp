from datetime import timedelta

import openpyxl
from django.utils import timezone
from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema

from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from rest_framework.generics import GenericAPIView
from rest_framework.permissions import IsAuthenticated
from io import BytesIO
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from api.models import Operation, Category, Target
from api.serializers import OperationSerializer

class OperationPDFView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OperationSerializer

    def get_queryset(self):
        days = int(self.request.query_params.get('days', 30))
        date_from = timezone.now() - timedelta(days=days)

        queryset = Operation.objects.filter(
            user=self.request.user,
            created_at__gte=date_from
        )

        operation_type = self.request.query_params.get('type')
        if operation_type:
            queryset = queryset.filter(type=operation_type)

        return queryset.select_related('categories').order_by('-created_at')

    @swagger_auto_schema(
        operation_id='Получение выписки в формате pdf',
        operation_description='Получение PDF файла с операциями и диаграммой',
        manual_parameters=[
            openapi.Parameter(
                name="type",
                in_=openapi.IN_QUERY,
                description="Фильтр по типу операций",
                type=openapi.TYPE_STRING,
                enum=["targets", "outcome", "income"],
                required=True,
            ),
            openapi.Parameter(
                name="days",
                in_=openapi.IN_QUERY,
                description="Количество дней для фильтрации",
                type=openapi.TYPE_INTEGER,
                required=True,
            ),
        ],
        responses={
            200: openapi.Response(description="PDF файл успешно получен"),
            500: openapi.Response(description="Ошибка генерации отчёта")
        }
    )
    def get(self, request):
        pdfmetrics.registerFont(TTFont('Bold', 'font/bold.ttf'))
        pdfmetrics.registerFont(TTFont('Regular', 'font/Regular.ttf'))
        operations = []
        total = 0.0
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        query = self.request.query_params.get('type')
        for operation in serializer.data:
            amount_str = str(operation['amount'])
            amount = float(amount_str.replace(' ', '').replace(',', '.'))
            total += amount

            formatted_amount = "{:,.2f}".format(amount).replace(",", " ")
            if query == "income" or query == "outcome":
                operations.append({
                    "category": Category.objects.get(id=operation["categories"]).name,
                    "amount": formatted_amount,
                    "date": operation["date"],
                    "_raw_amount": amount
                })
            elif query == "targets":
                operations.append({
                    "target": Target.objects.get(id=operation["target"]).name,
                    "amount": formatted_amount,
                    "date": operation["date"],
                    "_raw_amount": amount
                })
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4,
                                    leftMargin=2 * cm, rightMargin=2 * cm,
                                    topMargin=2 * cm, bottomMargin=2 * cm,
                                    encoding='utf-8')

            elements = []
            styles = getSampleStyleSheet()
            font_name = 'Regular'
            bold = "Bold"
            styles['Title'].fontName = font_name
            styles['Normal'].fontName = font_name

            title = Paragraph("Freenance", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 0.5 * cm))
            if query == "income" or query == "outcome":
                table_data = [
                    ["Date", "Category", "Amount"]
                ]
                for op in operations:
                    table_data.append([
                        op['date'],
                        op['category'],
                        op['amount']
                    ])
            elif query == "targets":
                table_data = [
                    ["Date", "Target", "Amount"]
                ]
                for op in operations:
                    table_data.append([
                        op['date'],
                        op['target'],
                        op['amount']
                    ])


            formatted_total = "{:,.2f}".format(total).replace(",", " ")
            table_data.append([
                "",
                "Total",
                formatted_total
            ])
            table = Table(table_data, colWidths=[6 * cm, 6 * cm, 6 * cm])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3A5FCD')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F0F8FF')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),

                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),

                ('FONTNAME', (1, -1), (3, -1), bold),
            ]))

            elements.append(table)

            doc.build(elements)
            buffer.seek(0)

            pdf_data = buffer.getvalue()
            if len(pdf_data) < 100:
                raise ValueError("Generated PDF is empty")

            response = HttpResponse(pdf_data, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="freenance_report.pdf"'
            return response

        except Exception as e:
            return HttpResponse(
                f"Error generating PDF: {str(e)}",
                status=500,
                content_type='text/plain'
            )


class OperationXLSView(GenericAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = OperationSerializer

    def get_queryset(self):
        days = int(self.request.query_params.get('days', 30))
        date_from = timezone.now() - timedelta(days=days)

        queryset = Operation.objects.filter(
            user=self.request.user,
            created_at__gte=date_from
        )

        operation_type = self.request.query_params.get('type')
        if operation_type:
            queryset = queryset.filter(type=operation_type)

        return queryset.select_related('categories').order_by('-created_at')

    @swagger_auto_schema(
        operation_id='Получение выписки в формате XLS',
        operation_description='Получение Excel файла с операциями',
        manual_parameters=[
            openapi.Parameter(
                name="type",
                in_=openapi.IN_QUERY,
                description="Фильтр по типу операций",
                type=openapi.TYPE_STRING,
                enum=["targets", "outcome", "income"],
                required=True,
            ),
            openapi.Parameter(
                name="days",
                in_=openapi.IN_QUERY,
                description="Количество дней для фильтрации",
                type=openapi.TYPE_INTEGER,
                required=True,
            ),
        ],
        responses={
            200: openapi.Response(description="XLS файл успешно получен"),
            500: openapi.Response(description="Ошибка генерации отчёта")
        }
    )
    def get(self, request):
        try:
            operations = []
            total = 0.0
            queryset = self.get_queryset()
            serializer = self.get_serializer(queryset, many=True)
            query = self.request.query_params.get('type')

            for operation in serializer.data:
                amount_str = str(operation['amount'])
                amount = float(amount_str.replace(' ', '').replace(',', '.'))
                total += amount

                formatted_amount = "{:,.2f}".format(amount).replace(",", " ")
                if query == "income" or query == "outcome":
                    operations.append({
                        "category": Category.objects.get(id=operation["categories"]).name,
                        "amount": formatted_amount,
                        "date": operation["date"],
                        "_raw_amount": amount
                    })
                elif query == "targets":
                    operations.append({
                        "target": Target.objects.get(id=operation["target"]).name,
                        "amount": formatted_amount,
                        "date": operation["date"],
                        "_raw_amount": amount
                    })

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Freenance Report"

            # Styles
            header_font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
            header_fill = PatternFill(start_color='3A5FCD', end_color='3A5FCD', fill_type='solid')
            normal_font = Font(name='Arial', size=11)
            total_font = Font(name='Arial', bold=True, size=11)
            alignment_center = Alignment(horizontal='center', vertical='center')
            alignment_left = Alignment(horizontal='left', vertical='center')
            alignment_right = Alignment(horizontal='right', vertical='center')
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            ws['A1'] = "Freenance"
            ws['A1'].font = Font(name='Arial', size=14)
            ws.merge_cells('A1:C1')
            ws.row_dimensions[1].height = 30

            if query == "income" or query == "outcome":
                headers = ["Date", "Category", "Amount"]
            elif query == "targets":
                headers = ["Date", "Target", "Amount"]

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment_center
                cell.border = thin_border

            for row_num, op in enumerate(operations, 4):
                if query == "income" or query == "outcome":
                    ws.cell(row=row_num, column=1, value=op['date']).alignment = alignment_center
                    ws.cell(row=row_num, column=2, value=op['category']).alignment = alignment_left
                    ws.cell(row=row_num, column=3, value=op['amount']).alignment = alignment_right
                elif query == "targets":
                    ws.cell(row=row_num, column=1, value=op['date']).alignment = alignment_center
                    ws.cell(row=row_num, column=2, value=op['target']).alignment = alignment_left
                    ws.cell(row=row_num, column=3, value=op['amount']).alignment = alignment_right

                for col_num in range(1, 4):
                    ws.cell(row=row_num, column=col_num).font = normal_font
                    ws.cell(row=row_num, column=col_num).border = thin_border
                    ws.cell(row=row_num, column=col_num).fill = PatternFill(
                        start_color='F0F8FF', end_color='F0F8FF', fill_type='solid')

            total_row = len(operations) + 4
            formatted_total = "{:,.2f}".format(total).replace(",", " ")

            ws.cell(row=total_row, column=1, value="").border = thin_border
            ws.cell(row=total_row, column=2, value="Total").font = total_font
            ws.cell(row=total_row, column=2).alignment = alignment_left
            ws.cell(row=total_row, column=2).border = thin_border
            ws.cell(row=total_row, column=3, value=formatted_total).font = total_font
            ws.cell(row=total_row, column=3).alignment = alignment_right
            ws.cell(row=total_row, column=3).border = thin_border

            for col_num in range(1, 4):
                ws.column_dimensions[get_column_letter(col_num)].width = 20

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="freenance_report.xlsx"'
            return response

        except Exception as e:
            return HttpResponse(
                f"Error generating XLS: {str(e)}",
                status=500,
                content_type='text/plain'
            )